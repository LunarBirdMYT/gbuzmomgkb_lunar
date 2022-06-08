from urllib.request import urlopen
# from urllib.parse import urlparse
# import wget
from bs4 import BeautifulSoup
# import re
from urllib.error import HTTPError, URLError
import openpyxl
from openpyxl import Workbook
import os


def to_txt(str):
    with open('text.txt', 'w') as file:
        file.write(str)

# with open('C:/Users/Оператор/Desktop/python/for_site_mgkb/Проверка.html', 'r', encoding='utf-8') as f:
#     # print(f)
#     html = f.read()  #.decode('utf-8')

#     soup = BeautifulSoup(html, 'html.parser')
#     # print(soup)

    
#     link_a = soup.find_all('a', href=True)
#     # print(len(link_tr))

# Передаем в функция номер строки - возвращает данные пациента            
def get_row_from_excel(number_row):
    try:
        book_in = openpyxl.load_workbook('эксель.xlsx')
        # book_in = openpyxl.open('эксель.xlsx', read_only=True)
        sheet = book_in.active
        code_region = sheet.cell(row=1+number_row, column=1).value
        region = sheet.cell(row=1+number_row, column=2).value
        data_for_check = sheet.cell(row=1+number_row, column=3).value
        data_after_check = sheet.cell(row=1+number_row, column=4).value
        name_tfoms = sheet.cell(row=1+number_row, column=5).value
        lenk_for_check = sheet.cell(row=1+number_row, column=6).value
        max_row = sheet.max_row
        # print('Максимальное количество строк: ', max_row)
        book_in.close()
        return tuple([code_region, region, data_for_check, data_after_check,
                      name_tfoms, lenk_for_check])
    except:
        print(input('Файл "эксель.xlsx" не найден\nСохраните файл в директории со скриптом\nДля продолжения нажмите Enter...'))
    


# print(get_row_from_excel(7))

# if os.path.exists('эксель.xlsx'):
#     try:
#         os.rename('эксель.xlsx', 'эксель1.xlsx')
#     except OSError:
#         print('Файл еще открыт')
with open('text.txt', 'w') as f:
    for i in range(1):
        
        print('<table border="0" class="tab" style="font-size: 12px; margin: 10px 0px 20px; border: 1px solid #000000; width: 1200px; color: #222222; font-family: arial, helvetica, sans-serif; background-color: #ffffff; border-collapse: collapse; height: 200px;">', file=f)
        
        for j in range(1):
            print('<thead style="text-align: center; border-left: 1px solid #000000; border-collapse: collapse;">', file=f)
            
            for k in range(1):
                print('<tr style="text-align: left; border: 1px solid #000000; border-collapse: collapse;">', file=f)
                
                for l in range(1):
                    data_zag = get_row_from_excel(0)
                    print('<th style="border-left: 1px solid #000000; border-collapse: collapse; padding: 7px 4px 7px 15px; height: 28px; width: 47px;">', file=f)
                    print(f'<span style="font-size: 18pt">{data_zag[0]}</span>', file=f)
                    print('</th>', file=f)
                    
                    print('<th style="border-left: 1px solid #000000; border-collapse: collapse; padding: 7px 4px 7px 15px; height: 28px; width: 117px;">', file=f)
                    print(f'<span style="font-size: 18pt">{data_zag[1]}</span>', file=f)
                    print('</th>', file=f)
                    
                    print('<th style="border-left: 1px solid #000000; border-collapse: collapse; padding: 7px 4px 7px 15px; height: 28px; width: 117px;">', file=f)
                    print(f'<span style="font-size: 18pt">{data_zag[2]}</span>', file=f)
                    print('</th>', file=f)
            
                    print('<th style="border-left: 1px solid #000000; border-collapse: collapse; padding: 7px 4px 7px 15px; height: 28px; width: 147px;">', file=f)
                    print(f'<span style="font-size: 18pt">{data_zag[3]}</span>', file=f)
                    print('</th>', file=f)
                    
                    print('<th style="padding: 7px 4px 7px 15px; color: #47433f; border-collapse: collapse; border: 1px solid #000000; height: 14px; width: 147px;">', file=f)
                    print(f'<span style="font-size: 18pt">{data_zag[4]}</span>', file=f)
                    print('</th>', file=f)
            
                print('</tr>', file=f)
            print('</thead>', file=f)
            
            
            
            # Часть за пустую строку
            print('<tr style="text-align: left; border: 1px solid #000000; border-collapse: collapse;">', file=f)
            
            print('<td style="padding: 7px 4px 7px 15px; color: #47433f; border-collapse: collapse; border: 1px solid #000000; height: 14px; width: 47px;">', file=f)
            print('<input placeholder="Поиск по коду" id="findCodeRegion" onkeyup="searchTableNumber()" type="text" maxlength="2">', file=f)
            print('</td>', file=f)
                    
            print('<td style="padding: 7px 4px 7px 15px; color: #47433f; border-collapse: collapse; border: 1px solid #000000; height: 14px; width: 117px;">', file=f)
            print('<input placeholder="Поиск по названию" id="findTitleRegion" onkeyup="searchTableTitle()" type="text" maxlength="70">', file=f)
            print('</td>', file=f)
                    
            print('<td style="padding: 7px 4px 7px 15px; color: #47433f; border-collapse: collapse; border: 1px solid #000000; height: 14px; width: 117px;">', file=f)
            print('</td>', file=f)
            
            print('<td style="padding: 7px 4px 7px 15px; color: #47433f; border-collapse: collapse; border: 1px solid #000000; height: 14px; width: 147px;">', file=f)
            print('</td>', file=f)
                    
            print('<td style="padding: 7px 4px 7px 15px; color: #47433f; border-collapse: collapse; border: 1px solid #000000; height: 14px; width: 147px;">', file=f)
            print('</td>', file=f)
            
            print('</tr>', file=f)
            print('<tbody id="myTable">', file=f)
            
            for m in range(1, 86):
                    print('<tr style="text-align: left; border: 1px solid #000000; border-collapse: collapse;">', file=f)
                
                    data_zag = get_row_from_excel(m)
                    print('<td style="padding: 7px 4px 7px 15px; color: #47433f; border-collapse: collapse; border: 1px solid #000000; height: 14px; width: 47px;">', file=f)
                    print(f'<span style="font-size: 18pt">{data_zag[0]}</span>', file=f)
                    print('</td>', file=f)
                    
                    print('<td style="padding: 7px 4px 7px 15px; color: #47433f; border-collapse: collapse; border: 1px solid #000000; height: 14px; width: 117px;">', file=f)
                    print(f'<span style="font-size: 18pt">{data_zag[1]}</span>', file=f)
                    print('</td>', file=f)
                    
                    # Условия для столбца необходимых данных для проверки
                    if 'функционал' in data_zag[2]:
                        print('<td style="padding: 7px 4px 7px 15px; color: #47433f; border-collapse: collapse; border: 1px solid #000000; height: 14px; width: 117px;">', file=f)
                        print(f'<span style="font-size: 15pt;color: salmon">{data_zag[2]}</span>', file=f)
                        print('</td>', file=f)
                    else:
                        print('<td style="padding: 7px 4px 7px 15px; color: #47433f; border-collapse: collapse; border: 1px solid #000000; height: 14px; width: 117px;">', file=f)
                        print(f'<span style="font-size: 18pt">{data_zag[2]}</span>', file=f)
                        print('</td>', file=f)
                    
                    # Условия для столбца возвращаемых данных
                    if data_zag[3] != None:
                        data_for_list = data_zag[3].split(', ')
                        count_li = len(data_for_list)
                        if data_zag[3] == 'раздел в разработке':
                            print('<td style="padding: 7px 4px 7px 15px; color: #47433f; border-collapse: collapse; border: 1px solid #000000; height: 14px; width: 147px;">', file=f)
                            print('<span style="font-size: 15pt;color: lightgrey">раздел в разработке</span>', file=f)
                            print('</td>', file=f)
                            
                        else:
                            print('<td style="padding: 7px 4px 7px 15px; color: #47433f; border-collapse: collapse; border: 1px solid #000000; height: 14px; width: 147px;">', file=f)
                            print('<ul style="padding: 7px 4px 7px 15px; font-size: 15pt">', file=f)
                            
                            for z in range(count_li):
                                print(f'<li>{data_for_list[z]}</li>', file=f)
                            # print(f'<span style="font-size: 18pt">{data_zag[3]}</span>', file=f)
                            print('</ul>', file=f)
                            print('</td>', file=f)
                    else:
                        print('<td style="padding: 7px 4px 7px 15px; color: #47433f; border-collapse: collapse; border: 1px solid #000000; height: 14px; width: 147px;">', file=f)
                        print('<span style="font-size: 18pt"></span>', file=f)
                        print('</td>', file=f)
                    
                    print('<td style="padding: 7px 4px 7px 15px; color: #47433f; border-collapse: collapse; border: 1px solid #000000; height: 14px; width: 147px;">', file=f)
                    print('<span style="font-size: 18pt">', file=f)
                    print(f'<a href="{data_zag[5]}" style="color: #27609b; text-decoration: none" target="_blank">{data_zag[4]}</a>', file=f)
                    print('</span>', file=f)
                    print('</td>', file=f)
                    print('</tr>', file=f)
            print('</tbody>', file=f)
        
        print('</table>', file=f)
        
        # Часть за скрипт
        print('<script>', file=f)
        print('function searchTableNumber() {', file=f)
        print('    var input, filter, found, table, tr, td, i, j;', file=f)
        print('    input = document.getElementById("findCodeRegion");', file=f)
        print('    filter = input.value.toUpperCase();', file=f)
        print('    table = document.getElementById("myTable");', file=f)
        print('    tr = table.getElementsByTagName("tr");', file=f)
        print('    for (i = 0; i < tr.length; i++) {', file=f)
        print('        td = tr[i].getElementsByTagName("td");', file=f)
        print('        for (j = 0; j < td.length; j++) {', file=f)
        print('            if (td[j].innerHTML.toUpperCase().indexOf(filter) > -1) {', file=f)
        print('                found = true;', file=f)
        print('            }', file=f)
        print('        }', file=f)
        print('        if (found) {', file=f)
        print('            tr[i].style.display = "";', file=f)
        print('            found = false;', file=f)
        print('        } else {', file=f)
        print('            tr[i].style.display = "none";', file=f)
        print('        }', file=f)
        print('    }', file=f)
        print('}', file=f)
        print('function searchTableTitle() {', file=f)
        print('    var input, filter, found, table, tr, td, i, j;', file=f)
        print('    input = document.getElementById("findTitleRegion");', file=f)
        print('    filter = input.value.toUpperCase();', file=f)
        print('    table = document.getElementById("myTable");', file=f)
        print('    tr = table.getElementsByTagName("tr");', file=f)
        print('    for (i = 0; i < tr.length; i++) {', file=f)
        print('        td = tr[i].getElementsByTagName("td");', file=f)
        print('        for (j = 0; j < td.length; j++) {', file=f)
        print('            if (td[j].innerHTML.toUpperCase().indexOf(filter) > -1) {', file=f)
        print('                found = true;', file=f)
        print('            }', file=f)
        print('        }', file=f)
        print('        if (found) {', file=f)
        print('            tr[i].style.display = "";', file=f)
        print('            found = false;', file=f)
        print('        } else {', file=f)
        print('            tr[i].style.display = "none";', file=f)
        print('        }', file=f)
        print('    }', file=f)
        print('}', file=f)
        print('</script>', file=f)

