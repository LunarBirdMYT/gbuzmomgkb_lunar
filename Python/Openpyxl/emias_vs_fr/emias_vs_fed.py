"""
UPD: 02/11/2021, v0.4
--- Внесено исправление для кода, на случай, когда с отчета bi присутствует
    тольно фамилия пациента
    
UPD: 20/01/2022 v0.5
--- Исправления конечной формы для ввода при помощи селениум

UPD: 22/02/2022 v0.6
--- Добавлен диагноз для выписки из ФР
    Исключен Отчет о переданных случаях в ФР в виду его долгой неработоспособности
"""


import openpyxl
from openpyxl import Workbook
# from openpyxl.styles import Font, Border, Side, Alignment

file_name_bi = 'Отчет .xlsx'
file_name_fr = 'Федеральный регистр.xlsx'


def try_except_func(list_from, search_string):
    try:
        search_index = list_from.index(search_string)
        return search_index
    except ValueError:
        print(input(f'Не найден столбец "{search_string}" в "{list_from}"'))


# Поскольку МИАЦ не починил отчет, то функция не используется
def err_from_EMIAS_stas():
    try:
        file_name_er = 'Отчет .xlsx'
        book  = openpyxl.open(file_name_er, read_only=True)
        sheet = book.active
    
        element = []
        for row in sheet.iter_rows(min_row=9, values_only=True):
            fio = (f'{row[2]} {row[3]} {row[4]}').upper()
            card = row[8]
            err = row[10]
            lst = tuple([fio, card, err])
            element.append(lst)
        return element
    except:
        print(input(f'Файл {file_name_er} не найден.'))
    finally:
        book.close()


def open_file(file_name):
    try:
        # Открываем книгу только для чтения.
        book = openpyxl.open(file_name, read_only=True)
        
        # Задаем активный лист, по умолчанию текущий.
        sheet = book.active
        
        # Создаем вложенные списки по всем элементам таблицы xlxs.
        element = []
        for row in sheet.iter_rows(values_only=True):
            element.append(list(row))
            
        # print(element)
        return element
    except Exception as e:
        print(f'Возникла ошибка:\n{e}')
        print(f'Вероятно: Файл с нужным именем не найден "{file_name}"!')
        print(f'  "{file_name}" с фед.регистра необходимо пересохранить!')
        print(input('  Переименуйте файл или проверьте формат "xlsx"!'))
    finally:
        # Закрываем книгу в любом случае
        book.close()


def open_earlier_file():
    try:
        file_name = 'ранее.xlsx'
        book  = openpyxl.open(file_name, read_only=True)
        sheet = book['Нет в ФР']
    
        element = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            element.append(row[0])
        return element
    except:
        print(input(f'Файл {file_name} не найден.'))
    finally:
        book.close()


# Функции по открытию фед.рег.
element_fr = open_file(file_name_fr)
del element_fr[0]
title_fr = element_fr[0]
del element_fr[0]

# Функция по открытию отчета с bi
element_bi = open_file(file_name_bi)
del element_bi[0]
title_bi = element_bi[0]
del element_bi[0:2]

# Индексы и проверки
ind_fio_fr = try_except_func(title_fr, 'ФИО')
ind_diagnosis_fr = try_except_func(title_fr, 'Диагноз')
ind_birthday_fr = try_except_func(title_fr, 'Дата рождения')
ind_f_bi = try_except_func(title_bi, 'Фамилия')  
ind_i_bi = try_except_func(title_bi, 'Имя')
ind_o_bi = try_except_func(title_bi, 'Отчество')
ind_date_hosp_bi = try_except_func(title_bi, 'Дата и время поступления')

# Функции для добавления пациентов в списки и сравнение по ФИО
fio_fr = dict()
for i in element_fr:
    if i[ind_fio_fr] != None:
        fio = i[ind_fio_fr].upper()
        diagnosis = i[ind_diagnosis_fr]
        bithday = i[ind_birthday_fr]
        fio_fr[fio] = tuple([diagnosis, bithday])

fio_bi = dict()
for j in element_bi:
    if j[ind_f_bi] and j[ind_i_bi] and j[ind_o_bi] != None:
        fio = j[ind_f_bi].strip() + " " + j[ind_i_bi].strip() + \
            " " + j[ind_o_bi].strip()
        date_hosp = j[ind_date_hosp_bi]
        fio_bi[fio.upper()] = date_hosp
    elif j[ind_f_bi] != None and j[ind_i_bi] != None and j[ind_o_bi] == None:
        fio = j[ind_f_bi].strip() + " " + j[ind_i_bi].strip()
        date_hosp = j[ind_date_hosp_bi]
        fio_bi[fio.upper()] = date_hosp
    elif j[ind_i_bi] == None and j[ind_o_bi] == None:
        fio = j[ind_f_bi].strip()
        date_hosp = j[ind_date_hosp_bi]
        fio_bi[fio.upper()] = date_hosp


# Список для вывода: есть в ФР, нет в ЕМИАС
fr_vs_emias = []
title_fr = ['ФИО', 'Фамилия', 'Имя' ,'Отчеcтво', 'Дата рождения','Диагноз']
fr_vs_emias.append(title_fr)

for i in fio_fr.keys():
    if i not in fio_bi.keys() and i != None:
        interim = []
        interim.append(i)
        iterim_1 = i.split()
        interim.append(iterim_1[0])
        interim.append(iterim_1[1])
        try:
            interim.append(iterim_1[2])
        except:
            interim.append(None)
        interim.append(fio_fr[i][1])  # Добавление диагноза
        interim.append(fio_fr[i][0])  # Добавление даты рождения
        
        fr_vs_emias.append(interim)
print('Нет в емиас всего: ',len(fr_vs_emias) - 1)


# Список для вывода: есть в ЕМИАС, нет в ФР
emias_vs_fr = []
title_bi = ['ФИО', 'Фамилия', 'Имя' ,'Отчеcтво', 'Дата госпитализации']
emias_vs_fr.append(title_bi)

# Проверка встречался ли он ранее
find_man_early = open_earlier_file()

for i, v in fio_bi.items():
        if i not in fio_fr.keys() and i != None:
            interim = []
            interim.append(i)
            iterim_1 = i.split()
            interim.append(iterim_1[0])
            try:
                interim.append(iterim_1[1])
            except:
                interim.append(None)
            try:
                interim.append(iterim_1[2])
            except:
                interim.append(None)
            interim.append(v)
                
            if i in find_man_early:  # Если был ранее
                interim.append('!!!Был ранее')  
            emias_vs_fr.append(interim)


# # Список с ошибками для дальнейшего сравнения
# err_from = err_from_EMIAS_stas()


# # Добавляем в файл случай стац и ошибку.        
# for i in err_from:
#     for j in emias_vs_fr[1:]:
#         if i[0] == j[0] and '"method":"UPDATE","status":"ok"}' not in i[2]:
#             if '"method":"CREATE","status":"ok"}' in i[2]:
#                 j.append(i[1])
#                 j.append(i[2])
#                 j.append('********ЭТОТ ПАРЯ УЖЕ В ФР*******')
#             else:
#                 j.append(i[1])
#                 j.append(i[2]) 
        
print('Нет в ФР всего: ',len(emias_vs_fr) - 1)


def to_exel():
    # Содание эксель файла
    wb = Workbook()
    wb['Sheet'].title = 'Нет в ЕМИАС'
    sh1 = wb.active

    # добавление данных в таблицу
    for i in fr_vs_emias:
        sh1.append(i)
        
    # Создаем новый лист
    sh2 = wb.create_sheet('Нет в ФР')
    # добавление данных в таблицу
    for i in emias_vs_fr:
        sh2.append(i)    
    
    # Создание книги и сохранение файла
    wb.save("Сравнение ФР с ЕМИАС и наоборот.xlsx")
    wb.close()

to_exel()

      
print()
print('Обработка файлов завершена успешно!')
print(input('_____  Нажмите любую клавишу...  _____'))    
      