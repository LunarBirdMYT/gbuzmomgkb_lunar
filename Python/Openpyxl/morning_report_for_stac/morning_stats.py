"""
update 20/11/2021:
-- Оптимизация в соответсвии с PEP-8
-- Изменения касательно вхадящих списков и элементов на более понятные
-- Переработаны во всех функциях добавление строк на вывод, исключен бредовый цикл

"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment

print('    ********* **********')
print('    Первая строка файла - Заголовки таблицы.')
print('    Со второй строки и далее данные по пациентам.')
print('    Имя исходного файла эксель:')
print('    Отчет по перечню пациентов (Стационар) ПО ОТДЕЛЕНИЯМ.xlsx')
print('    ********* **********')


def book_read():
    try:
        # Открываем книгу только для чтения.
        book_excel = openpyxl.open("Отчет по перечню пациентов (Стационар) ПО ОТДЕЛЕНИЯМ.xlsx", read_only=True)
        # Задаем активный лист, по умолчанию текущий.
        sheet = book_excel.active

        # Создаем вложенные списки по всем элементам таблицы xlxs.
        element = []
        for row in sheet.iter_rows(values_only=True):
            element.append(tuple(row))

        # print(element)

        return element
    except Exception as e:
        print(f'Ошибка: \n{e}')
        print('  Файл с нужным именем не найден!')
        print(input('  Переименуйте файл или проверьте формат "xlsx"!'))
    finally:
        # Закрываем книгу
        book_excel.close()


element = book_read()

# Вывод заголовков и присваивание индексов, а так же проверка вхождения
TITLE = element[0]
# for i in title:
#     print(title.index(i), i)


def try_except_value_error(lst, string):
    try:
        index_el = lst.index(string)
        return index_el
    except ValueError:
        print(input(f'Не найден столбец "{string}"'))


ind_otd = try_except_value_error(TITLE, 'Отделение')
ind_type_otd = try_except_value_error(TITLE, 'Тип отделения')
ind_time_in = try_except_value_error(TITLE, 'Время пребывания в стационаре, в днях')
ind_saturation = try_except_value_error(TITLE, 'Сатурация при динамическом наблюдении')
ind_dyn_status = try_except_value_error(TITLE, 'Состояние при последнем наблюдении')
age = try_except_value_error(TITLE, 'возраст')
temp_on_din = try_except_value_error(TITLE, 'Температура при динамическом наблюдении')
ivl = try_except_value_error(TITLE, 'ИВЛ')
oxygen = try_except_value_error(TITLE, 'Кислород')
# vaccine = try_except_ValueError(TITLE, 'Этап вакцинации')

# Удаляем заголовки для корректной работы цикла.
del element[0]


# Создаём функцию для стиля заголовков, ранее завдалась ширина файла
def sheet_title_aligment(sheet_for_aligment):
    dict_alig = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K',
                 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W',
                 'X', 'Y', 'Z']
    chr_aligment = '1'  # Чтобы получить стобец А1, В1 и далее

    try:
        # использую title[0], поскольку вложенные списки [[отд, фамилия, имя....]]
        for i in range(len(list(TITLE))):
            # Чтобы получить стобец А1, В1 и далее
            for_aligment = dict_alig[i] + chr_aligment
            sheet_for_aligment[for_aligment].alignment = Alignment(horizontal='distributed',
                                                                   vertical='top')
            # print('пошли фоглименты', for_aligment)

    except IndexError:
        print(input('Внимание! Количество заголовков превысило лимит \
                        Алфавита[dict_width] = 26! Исправьте скрипт!!!! \n \
                            Нажмите Enter, чтобы продолжить \n \
                                '))


# Ширина для столбцев после удаления
def sheet_del_width(sheet_for_del):
    # Заново собираем заголовки
    title_width_1 = []
    for rows in sheet_for_del.iter_rows(values_only=True, max_row=1):
        title_width_1.append(tuple(rows))
        pass
    title_width = title_width_1[0]
    # print(title_width)

    # Для выравния эксель столбцов, создаем словарь
    dict_width = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K',
                  'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W',
                  'X', 'Y', 'Z']
    # print(len(dict_width))
    # print(len(title_width))

    # А теперь по индексу эелемента в титуле, подбираем столбец и задаем ширину
    department = try_except_value_error(title_width, 'Отделение')
    surname = try_except_value_error(title_width, 'Фамилия')
    name = try_except_value_error(title_width, 'Имя')
    middle_name = try_except_value_error(title_width, 'Отчество')
    age_after_del = try_except_value_error(title_width, 'возраст')
    diagnosis = try_except_value_error(title_width, 'Диагноз по МКБ-10')
    data_arrival = try_except_value_error(title_width, 'Дата и время поступления')
    time_in_stacionar = try_except_value_error(title_width, 'Время пребывания в стационаре, в днях')
    status_final = try_except_value_error(title_width, 'Состояние при последнем наблюдении')
    satur_dinamic = try_except_value_error(title_width, 'Сатурация при динамическом наблюдении')
    ivl_ater_del = try_except_value_error(title_width, 'ИВЛ')
    oxygen_after_del = try_except_value_error(title_width, 'Кислород')
    date_out_plan = try_except_value_error(title_width, 'Планируемая дата выписки')
    temper_dynamic = try_except_value_error(title_width, 'Температура при динамическом наблюдении')

    # Задаем ширину
    sheet_for_del.column_dimensions[dict_width[department]].width = 58
    sheet_for_del.column_dimensions[dict_width[surname]].width = 21
    sheet_for_del.column_dimensions[dict_width[name]].width = 21
    sheet_for_del.column_dimensions[dict_width[middle_name]].width = 21
    sheet_for_del.column_dimensions[dict_width[age_after_del]].width = 10
    sheet_for_del.column_dimensions[dict_width[diagnosis]].width = 9
    sheet_for_del.column_dimensions[dict_width[data_arrival]].width = 17
    sheet_for_del.column_dimensions[dict_width[time_in_stacionar]].width = 10
    sheet_for_del.column_dimensions[dict_width[status_final]].width = 19
    sheet_for_del.column_dimensions[dict_width[satur_dinamic]].width = 11
    sheet_for_del.column_dimensions[dict_width[ivl_ater_del]].width = 10
    sheet_for_del.column_dimensions[dict_width[oxygen_after_del]].width = 10
    sheet_for_del.column_dimensions[dict_width[date_out_plan]].width = 16
    sheet_for_del.column_dimensions[dict_width[temper_dynamic]].width = 11


# Функции создания и первичной обработки книги.
def wb_open(name_list, title):
    # Создание книги
    wb = Workbook()
    wb['Sheet'].title = name_list
    sh1 = wb.active

    # Добавление заголовка.
    for i in title:
        sh1.append(i)

    # Стилизование заголовков
    sheet_title_aligment(sh1)

    # Добавление фильтра заголовка.
    sh1.auto_filter.ref = sh1.dimensions
    return wb, sh1


# Функции обработки стиля.
def sheet_style(sheet_in, count_column):

    # Индексы для корректного выравнивания столбцев и применения фильтров
    # +1 необходимо для корректного вычисления столбца, используются в sheet_style
    age_style = age + 1
    ind_time_in_style = ind_time_in + 1
    ind_saturation_style = ind_saturation + 1
    temp_on_din_style = temp_on_din + 1

    # Жирный заголовок.
    for i in range(count_column):
        sheet_in.cell(row=1, column=i+1).font = Font(bold=True)

    # Создание границы ячеек.
    max_row = sheet_in.max_row  # Количество строк
    for i in range(max_row):
        for j in range(count_column):
            sheet_in.cell(row=i+1, column=j+1).border = Border(left=Side(border_style='thin',
                                                               color='000000'),
                                                               right=Side(border_style='thin',
                                                               color='000000'),
                                                               top=Side(border_style='thin',
                                                               color='000000'),
                                                               bottom=Side(border_style='thin',
                                                               color='000000'),
                                                               )
            sheet_in.cell(row=i+2, column=age_style).alignment = Alignment(horizontal='center')
            sheet_in.cell(row=i+2, column=ind_time_in_style).alignment = Alignment(horizontal='center')
            sheet_in.cell(row=i+2, column=ind_saturation_style).alignment = Alignment(horizontal='center')
            sheet_in.cell(row=i+2, column=temp_on_din_style).alignment = Alignment(horizontal='center')
            sheet_in.cell(row=i+2, column=ind_time_in_style).font = Font(bold=True)
            # sheet.cell(row=i+2, column=(vaccine+1)).fill  = PatternFill(fill_type='solid',fgColor='FFC0CB')


# Функция закрытия и именования файла.
def wb_close(wb, name_file):
    wb.save(name_file)
    wb.close()


# Фукнция для удаления столбцов и выравнивания ширины
def del_column(sheet_del):
    # Заголовки для удаления(+1 необходим по тому что index считает с "0",
    # а при удалении с "1")
    # Проверок не проводится, ведь эти столбцы удаляются, и если их нет - ок.
    ind_type_otd_del = ind_type_otd + 1  # 13 столбец
    ind_come_in = TITLE.index('Самостоятельно поступил/По скорой помощи') + 1  # 12 столбец
    ind_sex = TITLE.index('пол') + 1  # 6 столбец
    # print('номер типа отделения', ind_type_otd_del)
    # print('номер поступления', ind_come_in)
    # print('номер пола', ind_sex)

    sheet_del.delete_cols(ind_type_otd_del)
    sheet_del.delete_cols(ind_come_in)
    sheet_del.delete_cols(ind_sex)
    sheet_del_width(sheet_del)


# Создаем циклы по необходимым параметрам.
# запись в файл более 10 дней
def days_10():
    name_list = '10 дней и более'
    wb, sh1 = wb_open(name_list, [TITLE])

    for el in element:
        if (el[ind_type_otd] == 'COVID-19' or el[ind_type_otd] == 'Прочие') and int(el[ind_time_in]) >= 10:
            sh1.append(el)

    # Обработка стилем.
    sheet_style(sh1, len(element[0]))
    del_column(sh1)
    sheet_del_width(sh1)

    # Присвоение имени и закрытие книги.
    name_file = 'Отчет по перечню пациентов (Стационар) 10 ДНЕЙ И БОЛЕЕ.xlsx'
    wb_close(wb, name_file)


# Запись в файл более 4 дней
def days_4():
    name_list = '4 дня и более'
    wb, sh1 = wb_open(name_list, [TITLE])

    # and int(el[ind_time_in]) >= 4) or ((el[ind_otd] == '1. Терапевтическое отделение №2 (Осташково)'
    # or el[ind_otd] == '1. Терапевтическое отделение №3 (Сухарево)') and int(el[ind_time_in]) >= 4):

    for el in element:
        if el[ind_type_otd] == 'Выздоравливающие':
            sh1.append(el)

    # Обработка стилем.
    sheet_style(sh1, len(element[0]))
    del_column(sh1)
    sheet_del_width(sh1)

    # Присвоение имени и закрытие книги.
    name_file = 'Количество пациентов на койках для выздоровления 4 И БОЛЕЕ ДНЕЙ.xlsx'
    wb_close(wb, name_file)


# Запись в файл без сатурации
def no_saturation():
    name_list = 'Без сатурации'
    wb, sh1 = wb_open(name_list, [TITLE])

    for el in element:
        if el[ind_saturation] is None:
            sh1.append(el)

    # Обработка стилем.
    sheet_style(sh1, len(element[0]))
    del_column(sh1)
    sheet_del_width(sh1)

    # Присвоение имени и закрытие книги.
    name_file = 'Отчет по перечню пациентов (Стационар) БЕЗ САТУРАЦИИ.xlsx'
    wb_close(wb, name_file)


# Запись в файл ср.тяжесть и сатурация более 95.
def saturation_up_95():
    name_list = 'Ср.тяж., сатур 95 и выше'
    wb, sh1 = wb_open(name_list, [TITLE])

    for el in element:
        if el[ind_saturation] is None:
            continue
        elif el[ind_dyn_status] == 'Средней тяжести' and int(el[ind_saturation]) >= 95:
            sh1.append(el)

    # Обработка стилем.
    sheet_style(sh1, len(element[0]))
    del_column(sh1)
    sheet_del_width(sh1)

    # Присвоение имени и закрытие книги.
    name_file = 'Отчет по перечню пациентов (Стационар) СР.ТЯЖЕСТЬ, САТУРАЦИЯ 95 И ВЫШЕ.xlsx'
    wb_close(wb, name_file)


def satisf_on_covid():
    name_list = 'Удов. сост. на ков койках'
    wb, sh1 = wb_open(name_list, [TITLE])

    for el in element:
        if el[ind_dyn_status] == 'Удовлетворительное' and el[ind_type_otd] == 'COVID-19':
            sh1.append(el)

    # Обработка стилем.
    sheet_style(sh1, len(element[0]))
    del_column(sh1)
    sheet_del_width(sh1)

    # Присвоение имени и закрытие книги.
    name_file = 'Отчет по перечню пациентов (Стационар) УДОВЛ. СОСТ. НА КОВИД.КОЙКАХ.xlsx'
    wb_close(wb, name_file)


def oxygen_10_days():
    name_list = '10 дней и более, без кислорода'
    wb, sh1 = wb_open(name_list, [TITLE])

    for el in element:
        if int(el[ind_time_in]) >= 10 and el[oxygen] != 'Да' and el[ivl] != 'Да':  # Более 10 дней
            sh1.append(el)

    # Обработка стилем.
    sheet_style(sh1, len(element[0]))
    del_column(sh1)
    sheet_del_width(sh1)

    # Присвоение имени и закрытие книги.
    name_file = 'Отчет по перечню пациентов (Стационар) БЕЗ ИВЛ И КИСЛОРОДА, БОЛЕЕ 10 ДНЕЙ.xlsx'
    wb_close(wb, name_file)

# def vaccine_func():
#     name_list = 'Поступившие и вакцинированные'
#     wb, sh1 = wb_open(name_list, [TITLE])

#     for el in element:
#         if el[vaccine] is not None:
#                sh1.append(el)

#     # Обработка стилем.
#     sheet_style(sh1, len(element[0]))
#     del_column(sh1)
#     sheet_del_width(sh1)
#
#     # Присвоение имени и закрытие книги.
#     name_file = 'Отчет по перечню пациентов (Стационар) ПОСТУПИВШИЕ И ВАКЦИНИРОВАННЫЕ.xlsx'
#     wb_close(wb, name_file)


# Вызовы функций формирования файлов
days_10()
days_4()
no_saturation()
saturation_up_95()
satisf_on_covid()
oxygen_10_days()
# vaccine_func()

print()
print('Обработка файлов завершена успешно!')
print(input('_____  Нажмите любую клавишу...  _____'))
